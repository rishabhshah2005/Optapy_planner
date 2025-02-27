# <add python path if required>

from domain import Lecture, TimeTable, generate_problem
from constraints import define_constraints
from optapy import get_class
import optapy.config
from optapy.types import Duration
from optapy import solver_factory_create
from optapy import score_manager_create
from openpyxl import load_workbook

solver_config = optapy.config.solver.SolverConfig() \
    .withEntityClasses(get_class(Lecture)) \
    .withSolutionClass(get_class(TimeTable)) \
    .withConstraintProviderClass(get_class(define_constraints)) \
    .withTerminationSpentLimit(Duration.ofSeconds(90))
    
solver = solver_factory_create(solver_config).buildSolver()
    
solution = solver.solve(generate_problem())
# print(solution)
score_manager = score_manager_create(solver_factory_create(solver_config))
score_explanation = score_manager.explainScore(solution)
# print(score_explanation)

def create_timetable_sheet(lst: list[Lecture]):
    template = load_workbook("template.xlsx")
    sheet1 = template['Sheet1']
    for l in lst:
        sheet1.cell(row=(l.timeslot.id+1), column=(int(l.division[1:])+2), value=l.subject+"-"+l.teacher.name+"-"+l.room.name)
    template.save("timetable.xlsx")
    print("Time table saved to timetable.xlsx")

create_timetable_sheet(solution.lecture_list)
